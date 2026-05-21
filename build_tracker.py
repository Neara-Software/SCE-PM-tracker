"""SCE Delivery Tracker — Neara Extract Edition.

Consumes a (distribution, transmission) extract pair and produces:
  - <out_dir>/SCE_delivery_tracker_<run_date>.xlsx (3 tabs)
  - <out_dir>/SM{n}_delivery.png  (one per scope month with any data)

See /Users/andrewwei/.claude/plans/tidy-meandering-lake.md for the spec.
"""

from __future__ import annotations

import argparse
import glob
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from typing import Iterable

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt


# --- Neara extract column lookup -------------------------------------------
# Header names (resolved at load time — column positions can shift between
# distribution/transmission and across daily extracts).

HEADER_SCOPE_MONTH = "Scope Month"
HEADER_CIRCUIT_NAME = "Circuit Name"
HEADER_CIRCUIT_ID = "Circuit Id"
HEADER_COMBO_KEY = "Combo Key"
HEADER_MILEAGE = "Mileage"
HEADER_MB_CHECK_DONE = "Gate Mb Check Done Ts R (epoch)"
HEADER_MB_FINAL_REVIEW_DONE = "Gate Mb Final Review Done Ts R (epoch)"
HEADER_PCQA_DONE = "Gate Pcqa Done Ts R (epoch)"
HEADER_SUBSTATION_PCQA_DONE = "Gate Substation Pcqa Done Ts R (epoch)"

REQUIRED_HEADERS = [
    HEADER_SCOPE_MONTH, HEADER_CIRCUIT_NAME, HEADER_CIRCUIT_ID,
    HEADER_COMBO_KEY, HEADER_MILEAGE,
    HEADER_MB_CHECK_DONE, HEADER_MB_FINAL_REVIEW_DONE,
    HEADER_PCQA_DONE, HEADER_SUBSTATION_PCQA_DONE,
]


def _resolve_indices(header_row):
    """Return {header_name: column_index}. First occurrence wins (Trans extracts have dup
    'Circuit Name' headers). Raises on missing required headers."""
    idx = {}
    for i, name in enumerate(header_row):
        if name and name not in idx:
            idx[name] = i
    missing = [h for h in REQUIRED_HEADERS if h not in idx]
    if missing:
        raise ValueError(
            "Neara extract missing required column(s): "
            + ", ".join(missing)
            + f". Headers found: {[h for h in header_row if h]}"
        )
    return idx


# --- Constants -------------------------------------------------------------

CUTOFF = date(2026, 1, 1)

SCOPE_MONTH_ORDER = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
SCOPE_MONTH_NUMBER = {m: i + 1 for i, m in enumerate(SCOPE_MONTH_ORDER)}
SCOPE_MONTH_ABBREV = {
    "jan": "January", "feb": "February", "mar": "March", "apr": "April",
    "may": "May", "jun": "June", "jul": "July", "aug": "August",
    "sep": "September", "oct": "October", "nov": "November", "dec": "December",
}

COLOUR_MB_CHECK = "#9b9b9b"        # grey
COLOUR_FINAL_REVIEW = "#1f6feb"    # blue
COLOUR_PCQA = "#2d6a4f"            # green
COLOUR_ACTIONABLE = "#a4193d"      # red
COLOUR_SCOPE_TOTAL = "#3d3d3d"     # charcoal

BUCKET_ORDER = [
    "MB Check done", "Final review done", "PC QA",
    "Actionable Total", "Scope Month Total",
]
BUCKET_COLOURS = {
    "MB Check done": COLOUR_MB_CHECK,
    "Final review done": COLOUR_FINAL_REVIEW,
    "PC QA": COLOUR_PCQA,
    "Actionable Total": COLOUR_ACTIONABLE,
    "Scope Month Total": COLOUR_SCOPE_TOTAL,
}


# --- Data model ------------------------------------------------------------


@dataclass
class Circuit:
    type: str
    scope_month: str
    name: str
    circuit_id: str
    mileage: float
    ready_date: date | None
    gis100_date: date | None
    pcqa_date: date | None


# --- Helpers ---------------------------------------------------------------


def epoch_to_date(value) -> date | None:
    """Decode an epoch-seconds value to a date, discarding pre-2026 dates."""
    if value is None or value == "" or value == 0:
        return None
    try:
        d = datetime.fromtimestamp(float(value), tz=timezone.utc).date()
    except (ValueError, OSError, OverflowError):
        return None
    if d < CUTOFF:
        return None
    return d


def canonical_scope_month(raw) -> str:
    """Normalise scope-month values like '03-Mar', '3', 'March', 'Mar' to 'March'."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    # Already canonical
    if s in SCOPE_MONTH_NUMBER:
        return s
    # Pure number
    try:
        n = int(float(s))
        if 1 <= n <= 12:
            return SCOPE_MONTH_ORDER[n - 1]
    except ValueError:
        pass
    # Common formats: '03-Mar', '3-Mar', 'Mar', 'Mar-03'
    parts = re.split(r"[\s\-_/]+", s)
    for p in parts:
        pl = p.lower()[:3]
        if pl in SCOPE_MONTH_ABBREV:
            return SCOPE_MONTH_ABBREV[pl]
        try:
            n = int(p)
            if 1 <= n <= 12:
                return SCOPE_MONTH_ORDER[n - 1]
        except ValueError:
            continue
    return s  # let it through unchanged so we can spot bad data


def load_monday_scope_totals(path: str) -> dict[str, float]:
    """Load mileage per scope month from a Monday Board extract, filtered to Neara CWA.

    Returns: { canonical_scope_month: total_miles }.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # Header row index — find the row containing 'Scope Month' and 'CWA - Total Mileage'.
    header_idx = None
    for i, row in enumerate(rows[:10]):
        if row and "Scope Month" in row and "CWA - Total Mileage" in row:
            header_idx = i
            break
    if header_idx is None:
        wb.close()
        raise ValueError(f"Could not find header row in {path}")
    header = rows[header_idx]
    idx = {h: i for i, h in enumerate(header) if h}
    sm_col = idx["Scope Month"]
    mi_col = idx["CWA - Total Mileage"]
    cwa_col = idx["Processing and Analysis CWA"]

    totals: dict[str, float] = defaultdict(float)
    for r in rows[header_idx + 1:]:
        if not r or not r[0]:
            continue
        cwa = r[cwa_col] if len(r) > cwa_col else None
        if not cwa or not str(cwa).startswith("2026_Neara"):
            continue
        sm_raw = r[sm_col] if len(r) > sm_col else None
        sm = canonical_scope_month(sm_raw)
        mi = r[mi_col] if len(r) > mi_col else 0
        try:
            mi = float(mi or 0)
        except (TypeError, ValueError):
            mi = 0.0
        totals[sm] += mi
    wb.close()
    return dict(totals)


def infer_type_from_filename(path: str) -> str:
    base = os.path.basename(path).lower()
    if "transmission" in base:
        return "Transmission"
    if "distribution" in base:
        return "Distribution"
    raise ValueError(f"Cannot infer type from filename: {path}")


# --- Extract loader --------------------------------------------------------


def load_extract(path: str, type_: str) -> list[Circuit]:
    """Read a Neara 'Circuit Tracking' extract into Circuit objects."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Circuit Tracking"] if "Circuit Tracking" in wb.sheetnames else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, ())
    idx = _resolve_indices(header)

    def cell(row, name):
        i = idx[name]
        return row[i] if i < len(row) else None

    circuits: list[Circuit] = []
    skipped_blank = 0
    unassigned = 0
    for row in rows_iter:
        if not row:
            continue
        circuit_id = cell(row, HEADER_CIRCUIT_ID)
        if not circuit_id or not str(circuit_id).strip():
            skipped_blank += 1
            continue
        combo_key = cell(row, HEADER_COMBO_KEY)
        if combo_key and str(combo_key).strip().lower() == "*not assigned*":
            unassigned += 1
            # No longer dropped — still count toward Actionable Total.

        mileage = cell(row, HEADER_MILEAGE) or 0
        try:
            mileage = float(mileage)
        except (TypeError, ValueError):
            mileage = 0.0

        ready = epoch_to_date(cell(row, HEADER_MB_CHECK_DONE))
        gis = epoch_to_date(cell(row, HEADER_MB_FINAL_REVIEW_DONE))
        pcqa_veg = epoch_to_date(cell(row, HEADER_PCQA_DONE))
        pcqa_sub = epoch_to_date(cell(row, HEADER_SUBSTATION_PCQA_DONE))
        pcqa = max(pcqa_veg, pcqa_sub) if (pcqa_veg and pcqa_sub) else None

        circuits.append(Circuit(
            type=type_,
            scope_month=canonical_scope_month(cell(row, HEADER_SCOPE_MONTH)),
            name=str(cell(row, HEADER_CIRCUIT_NAME) or "").strip(),
            circuit_id=str(circuit_id).strip(),
            mileage=mileage,
            ready_date=ready,
            gis100_date=gis,
            pcqa_date=pcqa,
        ))

    wb.close()
    print(f"  Loaded {len(circuits)} circuits from {os.path.basename(path)} "
          f"(skipped {skipped_blank} blank; included {unassigned} '*not assigned*')")
    return circuits


# --- Aggregations ----------------------------------------------------------


def _agg(circuits: Iterable[Circuit]) -> dict:
    total_n = total_mi = 0
    ready_n = ready_mi = 0
    gis_n = gis_mi = 0
    pcqa_n = pcqa_mi = 0
    for c in circuits:
        total_n += 1
        total_mi += c.mileage
        if c.ready_date:
            ready_n += 1
            ready_mi += c.mileage
        if c.gis100_date:
            gis_n += 1
            gis_mi += c.mileage
        if c.pcqa_date:
            pcqa_n += 1
            pcqa_mi += c.mileage
    return {
        "total_circuits": total_n, "total_miles": total_mi,
        "ready_circuits": ready_n, "ready_miles": ready_mi,
        "gis_circuits": gis_n, "gis_miles": gis_mi,
        "pcqa_circuits": pcqa_n, "pcqa_miles": pcqa_mi,
    }


def build_cumulative_summary(
    circuits: list[Circuit],
    scope_totals: dict[str, dict[str, float]] | None = None,
) -> list[dict]:
    """One dict per (type, scope_month) plus Combined rows.

    `scope_totals` is { type ('Distribution'|'Transmission'): { scope_month: miles } } from MB.
    """
    by_type_sm: dict[tuple[str, str], list[Circuit]] = defaultdict(list)
    by_sm: dict[str, list[Circuit]] = defaultdict(list)
    for c in circuits:
        by_type_sm[(c.type, c.scope_month)].append(c)
        by_sm[c.scope_month].append(c)

    scope_totals = scope_totals or {}

    def _scope_total(type_, sm):
        return scope_totals.get(type_, {}).get(sm, 0.0)

    scope_months = sorted(by_sm.keys(), key=lambda m: SCOPE_MONTH_NUMBER.get(m, 99))
    rows: list[dict] = []
    for sm in scope_months:
        d_agg = _agg(by_type_sm.get(("Distribution", sm), []))
        t_agg = _agg(by_type_sm.get(("Transmission", sm), []))
        d_scope = _scope_total("Distribution", sm)
        t_scope = _scope_total("Transmission", sm)
        rows.append({"Type": "Distribution", "Scope Month": sm,
                     "scope_total_miles": d_scope, **d_agg})
        rows.append({"Type": "Transmission", "Scope Month": sm,
                     "scope_total_miles": t_scope, **t_agg})
        c_agg = _agg(by_sm[sm])
        rows.append({"Type": "Combined", "Scope Month": sm,
                     "scope_total_miles": d_scope + t_scope, **c_agg})
    return rows


def build_daily_rate(circuits: list[Circuit]) -> list[dict]:
    """Long-format rows: (date, type, scope_month, bucket) with daily + cumulative miles."""
    bucket_field = {
        "MB Check done": "ready_date",
        "Final review done": "gis100_date",
        "PC QA": "pcqa_date",
    }
    daily: dict[tuple[date, str, str, str], list[float]] = defaultdict(lambda: [0.0, 0])
    for c in circuits:
        for bucket, field_name in bucket_field.items():
            d = getattr(c, field_name)
            if d:
                key = (d, c.type, c.scope_month, bucket)
                daily[key][0] += c.mileage
                daily[key][1] += 1

    rows = []
    cumulative: dict[tuple[str, str, str], float] = defaultdict(float)
    for key in sorted(daily.keys()):
        d, type_, sm, bucket = key
        miles, n = daily[key]
        cumulative[(type_, sm, bucket)] += miles
        rows.append({
            "Date": d,
            "Type": type_,
            "Scope Month": sm,
            "Bucket": bucket,
            "Miles": miles,
            "Circuits": n,
            "Cumulative Miles": cumulative[(type_, sm, bucket)],
        })
    return rows


# --- Workbook output -------------------------------------------------------


def _autosize(ws):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        length = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(length + 2, 40)


def write_workbook(cum_rows, daily_rows, circuits, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cumulative Progress Summary"

    header = [
        "Type", "Scope Month", "Total circuits", "Ready circuits",
        "GIS 100% circuits", "PC QA circuits",
        "Scope Month miles (MB)", "Actionable miles",
        "Ready miles", "GIS 100% miles", "PC QA miles",
        "Actionable % of Scope", "% Ready (of Actionable)",
        "% GIS 100% (of Actionable)", "% PC QA (of Actionable)",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    for r in cum_rows:
        actionable = r["total_miles"]
        scope = r.get("scope_total_miles", 0.0)
        denom = actionable or 1.0
        ws.append([
            r["Type"], r["Scope Month"], r["total_circuits"],
            r["ready_circuits"], r["gis_circuits"], r["pcqa_circuits"],
            round(scope, 3), round(actionable, 3),
            round(r["ready_miles"], 3), round(r["gis_miles"], 3),
            round(r["pcqa_miles"], 3),
            (actionable / scope) if scope else 0,
            r["ready_miles"] / denom if actionable else 0,
            r["gis_miles"] / denom if actionable else 0,
            r["pcqa_miles"] / denom if actionable else 0,
        ])
    for row in ws.iter_rows(min_row=2, min_col=12, max_col=15):
        for cell in row:
            cell.number_format = "0.0%"
    _autosize(ws)

    ws2 = wb.create_sheet("Daily Rate")
    ws2.append(["Date", "Type", "Scope Month", "Bucket", "Miles", "Circuits", "Cumulative Miles"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    for r in daily_rows:
        ws2.append([
            r["Date"], r["Type"], r["Scope Month"], r["Bucket"],
            round(r["Miles"], 3), r["Circuits"], round(r["Cumulative Miles"], 3),
        ])
    for row in ws2.iter_rows(min_row=2, max_col=1):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"
    _autosize(ws2)

    ws3 = wb.create_sheet("Per Circuit")
    ws3.append([
        "Type", "Scope Month", "Circuit Name", "Circuit ID", "Mileage",
        "MB Check done date", "Final review done date", "PC QA done date",
    ])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    for c in circuits:
        ws3.append([
            c.type, c.scope_month, c.name, c.circuit_id,
            round(c.mileage, 3),
            c.ready_date, c.gis100_date, c.pcqa_date,
        ])
    for row in ws3.iter_rows(min_row=2, min_col=6, max_col=8):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"
    _autosize(ws3)

    wb.save(out_path)


# --- Chart rendering -------------------------------------------------------


def _sm_summary(cum_rows, sm, type_):
    """Return the matching cumulative-summary dict for (sm, type_), or zeros."""
    for r in cum_rows:
        if r["Scope Month"] == sm and r["Type"] == type_:
            return r
    return {
        "total_miles": 0, "ready_miles": 0, "gis_miles": 0, "pcqa_miles": 0,
        "scope_total_miles": 0,
    }


def render_chart(sm: str, cum_rows: list[dict], out_path: str):
    trans = _sm_summary(cum_rows, sm, "Transmission")
    dist = _sm_summary(cum_rows, sm, "Distribution")

    sections = [
        ("Transmission", trans),
        ("Distribution", dist),
    ]

    bar_h = 0.65
    gap_within = 0.2
    gap_between = 1.4

    fig_w = 11
    fig_h = 6.5
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))

    cur = 0.0
    section_centers = []

    section_data = []
    for sect_label, data in sections:
        actionable = data["total_miles"]
        scope_total = data.get("scope_total_miles", 0)
        bars = [
            ("MB Check done", data["ready_miles"], COLOUR_MB_CHECK),
            ("Final review done", data["gis_miles"], COLOUR_FINAL_REVIEW),
            ("PC QA", data["pcqa_miles"], COLOUR_PCQA),
            ("Actionable Total", actionable, COLOUR_ACTIONABLE),
            ("Scope Month Total", scope_total, COLOUR_SCOPE_TOTAL),
        ]
        ys = []
        for _ in bars:
            cur -= (bar_h + gap_within)
            ys.append(cur)
        section_centers.append((sect_label, sum(ys) / len(ys)))
        section_data.append((sect_label, actionable, scope_total, bars, ys))
        cur -= gap_between

    max_total = max(
        (max(d[1], d[2]) for d in section_data), default=1
    ) or 1
    x_pad = max_total * 0.20

    for sect_label, actionable, scope_total, bars, ys in section_data:
        for (label, miles, colour), y in zip(bars, ys):
            ax.barh(y, miles, height=bar_h, color=colour, edgecolor="none")
            # % label inside coloured progress bars (against Scope Month Total denominator)
            if label in ("MB Check done", "Final review done", "PC QA") and scope_total > 0:
                pct = miles / scope_total
                if miles > max_total * 0.08:
                    ax.text(
                        miles / 2, y, f"{pct:.0%}",
                        va="center", ha="center",
                        color="white", fontsize=10, fontweight="bold",
                    )
            # Miles label at bar end
            if miles > 0:
                ax.text(
                    miles + max_total * 0.01, y,
                    f"{miles:,.0f}",
                    va="center", ha="left", fontsize=10,
                )

    ax.set_yticks([c for _, c in section_centers])
    ax.set_yticklabels([s for s, _ in section_centers], fontsize=12)
    ax.set_xlim(0, max_total + x_pad)
    ax.set_xticks([])
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["bottom"].set_visible(False)
    ax.tick_params(axis="y", which="both", left=False)

    sm_num = SCOPE_MONTH_NUMBER.get(sm, "?")
    # Title + subtitle with explicit pad above the chart area
    ax.set_title(
        f"Scope Month {sm_num} delivery",
        loc="left", pad=42, fontsize=20, color="#2d6a4f", fontweight="bold",
    )
    ax.text(
        0.0, 1.02, "(miles)",
        transform=ax.transAxes, ha="left", va="bottom",
        fontsize=11, color="#666",
    )

    handles = [
        plt.Rectangle((0, 0), 1, 1, color=BUCKET_COLOURS[b])
        for b in BUCKET_ORDER
    ]
    ax.legend(
        handles, BUCKET_ORDER,
        loc="lower center", bbox_to_anchor=(0.5, -0.16),
        ncol=5, frameon=False, fontsize=9,
    )

    plt.subplots_adjust(left=0.14, right=0.95, top=0.78, bottom=0.16)
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


# --- Entry point -----------------------------------------------------------


def _filter_real_xlsx(paths):
    """Drop Excel lock-files (start with '~$') and non-existent paths."""
    return [p for p in paths if not os.path.basename(p).startswith("~$")]


def _autodetect_pair():
    """Find the newest Neara extract pair under inputs/<dated>/, then fall back to Example_files."""
    here = os.path.dirname(os.path.abspath(__file__))
    # Preferred: inputs/<dated>/  (newest first)
    dated_dirs = sorted(
        (d for d in glob.glob(os.path.join(here, "inputs", "*")) if os.path.isdir(d)),
        reverse=True,
    )
    fallback = [os.path.join(here, "..", "Non_code_documents", "SCE_Tracker", "Example_files")]
    for d in dated_dirs + fallback:
        dist = _filter_real_xlsx(sorted(glob.glob(os.path.join(d, "*-distribution.xlsx"))))
        trans = _filter_real_xlsx(sorted(glob.glob(os.path.join(d, "*-transmission.xlsx"))))
        if dist and trans:
            return dist[-1], trans[-1]
    return None, None


def _autodetect_mb_pair(neara_dist_path: str):
    """Find Monday Board extracts in the same folder as the Neara distribution extract."""
    folder = os.path.dirname(neara_dist_path)
    dist = _filter_real_xlsx(sorted(glob.glob(os.path.join(folder, "*VM_LiDAR*Distribution*.xlsx"))))
    trans = _filter_real_xlsx(sorted(glob.glob(os.path.join(folder, "*VM_LiDAR*Transmission*.xlsx"))))
    return (dist[-1] if dist else None, trans[-1] if trans else None)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--distribution", help="Neara *-distribution.xlsx")
    parser.add_argument("--transmission", help="Neara *-transmission.xlsx")
    parser.add_argument("--mb-distribution", help="Monday Board Distribution .xlsx (for scope totals)")
    parser.add_argument("--mb-transmission", help="Monday Board Transmission .xlsx (for scope totals)")
    parser.add_argument("--out-dir", help="Output dir (default: ./outputs/YYYYMMDD/)")
    parser.add_argument("--scope-months", help="Comma list of months to chart (default: all)")
    args = parser.parse_args()

    dist_path, trans_path = args.distribution, args.transmission
    if not dist_path or not trans_path:
        auto_d, auto_t = _autodetect_pair()
        dist_path = dist_path or auto_d
        trans_path = trans_path or auto_t
    if not dist_path or not trans_path:
        parser.error("Provide --distribution and --transmission (or place dated extracts in inputs/).")

    mb_dist = args.mb_distribution
    mb_trans = args.mb_transmission
    if not mb_dist or not mb_trans:
        auto_md, auto_mt = _autodetect_mb_pair(dist_path)
        mb_dist = mb_dist or auto_md
        mb_trans = mb_trans or auto_mt

    run_date = datetime.now().strftime("%Y%m%d")
    out_dir = args.out_dir or os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "outputs", run_date
    )
    os.makedirs(out_dir, exist_ok=True)

    print(f"Distribution:    {dist_path}")
    print(f"Transmission:    {trans_path}")
    print(f"MB Distribution: {mb_dist or '(none — scope totals will be 0)'}")
    print(f"MB Transmission: {mb_trans or '(none — scope totals will be 0)'}")
    print(f"Output dir:      {out_dir}")
    print()

    print("Loading Neara extracts...")
    circuits = load_extract(dist_path, "Distribution") + load_extract(trans_path, "Transmission")
    print(f"  Total circuits in scope: {len(circuits)}")

    scope_totals: dict[str, dict[str, float]] = {"Distribution": {}, "Transmission": {}}
    if mb_dist:
        print(f"\nLoading MB Distribution scope totals from {os.path.basename(mb_dist)}...")
        scope_totals["Distribution"] = load_monday_scope_totals(mb_dist)
        print(f"  {sum(scope_totals['Distribution'].values()):,.1f} mi across "
              f"{len(scope_totals['Distribution'])} scope month(s) (Neara CWA)")
    if mb_trans:
        print(f"Loading MB Transmission scope totals from {os.path.basename(mb_trans)}...")
        scope_totals["Transmission"] = load_monday_scope_totals(mb_trans)
        print(f"  {sum(scope_totals['Transmission'].values()):,.1f} mi across "
              f"{len(scope_totals['Transmission'])} scope month(s) (Neara CWA)")

    print("\nBuilding aggregations...")
    cum_rows = build_cumulative_summary(circuits, scope_totals)
    daily_rows = build_daily_rate(circuits)

    workbook_path = os.path.join(out_dir, f"SCE_delivery_tracker_{run_date}.xlsx")
    print(f"\nWriting workbook: {workbook_path}")
    write_workbook(cum_rows, daily_rows, circuits, workbook_path)

    scope_months_present = sorted(
        {r["Scope Month"] for r in cum_rows if r["Type"] == "Combined"},
        key=lambda m: SCOPE_MONTH_NUMBER.get(m, 99),
    )
    if args.scope_months:
        wanted_nums = {int(x) for x in args.scope_months.split(",") if x.strip()}
        scope_months_present = [
            m for m in scope_months_present if SCOPE_MONTH_NUMBER.get(m) in wanted_nums
        ]

    print(f"\nRendering charts for {len(scope_months_present)} scope month(s)...")
    for sm in scope_months_present:
        n = SCOPE_MONTH_NUMBER.get(sm, "X")
        png = os.path.join(out_dir, f"SM{n}_delivery.png")
        render_chart(sm, cum_rows, png)
        print(f"  {png}")

    print("\nDone.")


if __name__ == "__main__":
    main()
